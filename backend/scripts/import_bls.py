"""
Builds the occupations dataset that backs Career Outlook and Job Market
Demand: for each of Fork's majors, which occupations does the federal
CIP-to-SOC crosswalk connect it to, and what does BLS say about each one.

Three raw federal files feed this, none of them committed to the repo (see
.gitignore) because they're large and freely re-downloadable:

  1. CIP-SOC Crosswalk (NCES/BLS, 2020 CIP / 2018 SOC)
     https://nces.ed.gov/ipeds/cipcode/Files/CIP2020_SOC2018_Crosswalk.xlsx
     Which occupations a field of study is connected to. This is EXPERT
     JUDGMENT, not outcomes data — it says "this program is commonly
     related to this occupation," not "graduates of this program actually
     work here." That distinction is load-bearing and ships as its own
     limitation, separate from the College Scorecard one, because the two
     datasets are weak in different ways.

  2. OEWS National, May 2025 (BLS)
     https://www.bls.gov/oes/tables.htm -> National -> May 2025
     National median wage per occupation. This is the newest wage figure
     available, newer than the wage column bundled into file 3 below — so
     wage comes from here, not from the projections file, even though the
     projections file has its own (older) wage column.

  3. Occupational Projections Data, 2024-2034 (BLS Employment Projections)
     https://www.bls.gov/emp/data/occupational-data.htm
     Growth rate, annual job openings, and typical entry-level education
     per occupation. OEWS doesn't have any of these, so they come from here
     instead.

Each fact keeps the source and date of the file it actually came from.
Wage is dated May 2025; growth/openings/education are dated the 2024-2034
cycle. Bundling them under one source label would misrepresent how current
each individual figure is.
"""

import csv
import json
import os
import sys
from datetime import date

import openpyxl

INSTITUTION_ID = "unt"

CROSSWALK_SOURCE = (
    "U.S. Dept. of Education (NCES) / U.S. Dept. of Labor (BLS), "
    "2020 CIP to 2018 SOC Crosswalk"
)
CROSSWALK_URL = "https://nces.ed.gov/ipeds/cipcode/Files/CIP2020_SOC2018_Crosswalk.xlsx"
CROSSWALK_LIMITATION = (
    "Which occupations connect to this field of study comes from a federal "
    "crosswalk built on expert judgment about the skills a program and an "
    "occupation have in common. It does NOT show where this program's "
    "actual graduates went to work, does not represent placement "
    "outcomes, and does not guarantee that this degree leads to any "
    "occupation listed here."
)

WAGE_SOURCE = "U.S. Dept. of Labor (BLS), Occupational Employment and Wage Statistics (OEWS), National"
WAGE_RELEASE = "May 2025"

PROJECTIONS_SOURCE = "U.S. Dept. of Labor (BLS), Employment Projections"
PROJECTIONS_CYCLE = "2024-2034"

# Fork major key -> the exact 2020 CIP title the crosswalk uses. Matched by
# title text, not by numeric prefix: our 4-digit Scorecard CIP codes each
# roll up several distinct 6-digit crosswalk entries with DIFFERENT
# occupation lists (e.g. 11.01 alone covers six different computing
# sub-programs), so matching by prefix would silently blend unrelated
# programs together. Matching the exact title that Scorecard itself
# reports for the field keeps this to the one 6-digit code that's
# genuinely the same program.
MAJOR_CIP_TITLE = {
    "computer_science": "Computer and Information Sciences, General.",
    "information_technology": "Computer and Information Sciences, General.",
    "business_administration": "Business Administration and Management, General.",
    "psychology_ba": "Psychology, General.",
    "psychology_bs": "Psychology, General.",
    "mechanical_energy_engineering": "Mechanical Engineering.",
}


def _clean_soc_code(raw) -> str:
    """CIP-SOC crosswalk SOC codes are plain strings like '15-1211'."""
    return str(raw).strip()


def _clean_occ_code_from_projections(raw: str) -> str:
    """The Employment Projections CSV export wraps SOC codes as an Excel
    formula to stop spreadsheet apps from mangling the dash: ="15-1211".
    Strip that back to a plain code."""
    return raw.strip().strip("=").strip('"')


def _clean_title(raw: str) -> str:
    """Occupation titles in the projections file carry a pile of alternate
    job titles appended after asterisks, meant for keyword search on
    BLS's own site — not for display here."""
    return raw.split("*")[0].strip()


def _num(raw):
    """Parses a BLS number field. Returns None for anything that isn't a
    real number (blanks, 'N/A', suppression markers) rather than guessing —
    same rule as the Scorecard importer."""
    if raw is None:
        return None
    v = str(raw).strip().replace(",", "")
    if v in ("", "N/A", "#", "*", "**"):
        return None
    try:
        return float(v)
    except ValueError:
        return None


def load_crosswalk(path: str) -> dict:
    """title -> list of {soc_code, soc_title}, for every occupation the
    crosswalk connects to that exact CIP title."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["CIP-SOC"]
    rows = ws.iter_rows(min_row=2, values_only=True)

    by_title: dict[str, list[dict]] = {}
    for cip_code, cip_title, soc_code, soc_title in rows:
        if not cip_title:
            continue
        by_title.setdefault(str(cip_title).strip(), []).append(
            {"soc_code": _clean_soc_code(soc_code), "soc_title": str(soc_title).strip()}
        )
    return by_title


def load_wages(path: str) -> dict:
    """SOC code -> {median_annual_wage, employment}, national, all
    industries. Only 'detailed' rows (single occupations) — the file also
    has 'major'/'minor'/'broad' rollup rows we don't want."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {h: i for i, h in enumerate(headers)}

    wages = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["O_GROUP"]] != "detailed":
            continue
        code = row[idx["OCC_CODE"]]
        wages[code] = {
            "median_annual_wage": _num(row[idx["A_MEDIAN"]]),
            "employment": _num(row[idx["TOT_EMP"]]),
        }
    return wages


def load_projections(path: str) -> dict:
    """SOC code -> growth/openings/education, from the 2024-2034 cycle."""
    projections = {}
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            code = _clean_occ_code_from_projections(row["Occupation Code"])
            projections[code] = {
                "title": _clean_title(row["Occupation Title"]),
                "percent_change": _num(row["Employment Percent Change, 2024-2034"]),
                "annual_openings": _num(
                    row["Occupational Openings, 2024-2034 Annual Average"]
                ),
                "typical_education": (row["Typical Entry-Level Education"] or "").strip()
                or None,
            }
    return projections


def build(crosswalk_path: str, wages_path: str, projections_path: str) -> dict:
    crosswalk = load_crosswalk(crosswalk_path)
    wages = load_wages(wages_path)
    projections = load_projections(projections_path)

    majors = {}
    for major_key, cip_title in MAJOR_CIP_TITLE.items():
        matches = crosswalk.get(cip_title)
        if not matches:
            raise SystemExit(
                f"No crosswalk entries found for CIP title {cip_title!r} "
                f"(major '{major_key}'). Either the crosswalk file changed "
                "or this title doesn't match it exactly."
            )

        occupations = []
        for m in matches:
            soc = m["soc_code"]
            wage = wages.get(soc, {})
            proj = projections.get(soc, {})
            occupations.append(
                {
                    "soc_code": soc,
                    "title": m["soc_title"],
                    "median_annual_wage": wage.get("median_annual_wage"),
                    "national_employment": wage.get("employment"),
                    "percent_change_2024_2034": proj.get("percent_change"),
                    "annual_openings": proj.get("annual_openings"),
                    "typical_education": proj.get("typical_education"),
                }
            )

        # Sorted by national employment, descending — an objective BLS
        # figure, not a judgment call about which jobs matter. Missing
        # employment figures sort last rather than crashing or defaulting
        # to 0 (which would wrongly rank them dead last as if employment
        # were known to be zero — it's just unreported).
        occupations.sort(
            key=lambda o: (
                o["national_employment"] is None,
                -(o["national_employment"] or 0),
            )
        )

        majors[major_key] = {
            "cip_title": cip_title,
            "occupations": occupations,
        }

    return {
        "_README": (
            "Occupations connected to each major, built by "
            "scripts/import_bls.py from three federal files: the NCES/BLS "
            "CIP-SOC crosswalk (which occupations), OEWS May 2025 (wage), "
            "and BLS Employment Projections 2024-2034 (growth, openings, "
            "typical education). Do not hand-edit — re-run the script."
        ),
        "institution_id": INSTITUTION_ID,
        "crosswalk_source": CROSSWALK_SOURCE,
        "crosswalk_source_url": CROSSWALK_URL,
        "crosswalk_limitation": CROSSWALK_LIMITATION,
        "wage_source": WAGE_SOURCE,
        "wage_release": WAGE_RELEASE,
        "projections_source": PROJECTIONS_SOURCE,
        "projections_cycle": PROJECTIONS_CYCLE,
        "retrieved": date.today().isoformat(),
        "majors": majors,
    }


def main():
    if len(sys.argv) < 4:
        raise SystemExit(
            "usage: python scripts/import_bls.py "
            "<CIP2020_SOC2018_Crosswalk.xlsx> <oesm25nat/national_M2025_dl.xlsx> "
            "<Employment_Projections.csv>"
        )
    crosswalk_path, wages_path, projections_path = sys.argv[1:4]
    for p in (crosswalk_path, wages_path, projections_path):
        if not os.path.exists(p):
            raise SystemExit(f"No such file: {p}")

    data = build(crosswalk_path, wages_path, projections_path)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "data_sources", "bls")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{INSTITUTION_ID}_occupations.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
        f.write("\n")

    print(f"Wrote {out_path}")
    for key, major in data["majors"].items():
        print(f"  {key}: {len(major['occupations'])} occupations")
        for occ in major["occupations"][:3]:
            wage = (
                f"${occ['median_annual_wage']:,.0f}"
                if occ["median_annual_wage"] is not None
                else "no wage"
            )
            emp = (
                f"{occ['national_employment']:,.0f} employed"
                if occ["national_employment"] is not None
                else "employment unknown"
            )
            print(f"    {occ['title']:<45} {wage:<12} {emp}")


if __name__ == "__main__":
    main()
